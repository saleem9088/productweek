import openpyxl


class Test_Data:

    def getTestData(self, test_name):
        try:
            dataDict = {}
            book = openpyxl.load_workbook("../Test_Data.xlsx")
            sheet = book.active
            for i in range(1, sheet.max_row + 1):  # to get rows
                if sheet.cell(row=i, column=1).value == test_name:

                    for j in range(2, sheet.max_column + 1):  # to get columns
                        dataDict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            return dataDict
        except FileNotFoundError as exp:
            print("Exception occurred while performing fetching values from excel", exp)
