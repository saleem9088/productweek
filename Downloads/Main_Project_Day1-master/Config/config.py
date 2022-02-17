import openpyxl
import string
import random


class Test_Data:
    # print(cell.value)
    # sheet.cell(row=2, column=2).value = "Rahul"
    # print(sheet.cell(row=2, column=2).value)
    # print(sheet.max_row)
    # print(sheet.max_column)
    # print(sheet['A5'].value)

    def getTestData(self, data_section,testcase_name):
        try:
            data_dict = {}
            book = openpyxl.load_workbook("../Config/Test_Data_List.xlsx")
            #sheet = book.active
            sheet = book.get_sheet_by_name(data_section)
            for i in range(1, sheet.max_row + 1):  # to get rows
                if sheet.cell(row=i, column=1).value == testcase_name:

                    for j in range(2, sheet.max_column + 1):  # to get columns
                        data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            return data_dict
        except FileNotFoundError as e:
            print("Exception occurred while performing file operation", e)

    def create_random_text(self):
        S = 3
        ran = ''.join(random.choices(string.ascii_uppercase + string.digits, k=S))
        print("The randomly generated string is : " + str(ran))
        ran = str(ran)
        print(ran)
        return ran

